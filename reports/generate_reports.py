#!/usr/bin/env python3
"""Генерация отчётов по сообщениям об аренде мотобайков."""

import json
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

DB_PATH = Path(__file__).parent.parent / "data" / "tg_search.db"
REPORTS_DIR = Path(__file__).parent

SQL = """
SELECT m.message_id, m.text, m.date, c.title, c.username
FROM messages m
INNER JOIN (
  SELECT rowid FROM messages_fts
  WHERE messages_fts MATCH '(байк OR мотобайк OR мотоцикл OR скутер) AND (аренда OR арендовать OR снять OR взять)'
) fts ON m.id = fts.rowid
LEFT JOIN channels c ON m.channel_id = c.channel_id
WHERE LENGTH(m.text) < 400
  AND m.text NOT LIKE '%t.me/%'
  AND m.text NOT LIKE '%бат за сутки%'
  AND m.text NOT LIKE '%Подберём%'
  AND m.text NOT LIKE '%ЗАКАЗАТЬ%'
  AND m.text NOT LIKE '%Предлагает в аренду%'
  AND m.text NOT LIKE '%Co.,Ltd%'
  AND m.text NOT LIKE '%@Motors%'
  AND m.text NOT LIKE '%Номер для связи%'
ORDER BY m.date DESC;
"""


def fetch_messages():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(SQL).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def parse_date(date_val):
    if isinstance(date_val, (int, float)):
        return datetime.fromtimestamp(date_val, tz=timezone.utc)
    if isinstance(date_val, str):
        # fromisoformat поддерживает +00:00 (Python 3.7+)
        try:
            dt = datetime.fromisoformat(date_val)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
            try:
                return datetime.strptime(date_val, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


def build_message_link(username, message_id):
    if username:
        return f"https://t.me/{username}/{message_id}"
    return ""


def generate_chart_html(messages):
    by_day = defaultdict(int)
    by_week = defaultdict(int)
    by_month = defaultdict(int)

    for msg in messages:
        dt = parse_date(msg["date"])
        if not dt:
            continue
        by_day[dt.strftime("%Y-%m-%d")] += 1
        # ISO week: YYYY-Www
        by_week[dt.strftime("%G-W%V")] += 1
        by_month[dt.strftime("%Y-%m")] += 1

    def sorted_labels_values(d):
        items = sorted(d.items())
        return [k for k, _ in items], [v for _, v in items]

    day_labels, day_values = sorted_labels_values(by_day)
    week_labels, week_values = sorted_labels_values(by_week)
    month_labels, month_values = sorted_labels_values(by_month)

    total = len(messages)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Аренда мотобайков — динамика сообщений</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; color: #333; }}
  .container {{ max-width: 960px; margin: 40px auto; padding: 0 20px; }}
  h1 {{ font-size: 1.5rem; margin-bottom: 6px; }}
  .subtitle {{ color: #666; font-size: 0.9rem; margin-bottom: 24px; }}
  .controls {{ display: flex; gap: 8px; margin-bottom: 20px; }}
  .btn {{ padding: 8px 20px; border: 2px solid #4a90d9; background: white; color: #4a90d9;
          border-radius: 6px; cursor: pointer; font-size: 0.9rem; font-weight: 500; transition: all .15s; }}
  .btn.active, .btn:hover {{ background: #4a90d9; color: white; }}
  .chart-wrap {{ background: white; border-radius: 10px; padding: 24px; box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  canvas {{ max-height: 420px; }}
</style>
</head>
<body>
<div class="container">
  <h1>Аренда мотобайков / скутеров — Паттайя</h1>
  <p class="subtitle">Всего сообщений: <strong>{total}</strong> &nbsp;|&nbsp; Данные из Telegram-каналов</p>
  <div class="controls">
    <button class="btn active" onclick="setMode('day', this)">По дням</button>
    <button class="btn" onclick="setMode('week', this)">По неделям</button>
    <button class="btn" onclick="setMode('month', this)">По месяцам</button>
  </div>
  <div class="chart-wrap">
    <canvas id="chart"></canvas>
  </div>
</div>
<script>
const DATA = {{
  day:   {{ labels: {json.dumps(day_labels, ensure_ascii=False)},   values: {json.dumps(day_values)} }},
  week:  {{ labels: {json.dumps(week_labels, ensure_ascii=False)},  values: {json.dumps(week_values)} }},
  month: {{ labels: {json.dumps(month_labels, ensure_ascii=False)}, values: {json.dumps(month_values)} }}
}};

const ctx = document.getElementById('chart').getContext('2d');
const chart = new Chart(ctx, {{
  type: 'bar',
  data: {{
    labels: DATA.day.labels,
    datasets: [{{
      label: 'Сообщений',
      data: DATA.day.values,
      backgroundColor: 'rgba(74,144,217,0.7)',
      borderColor: 'rgba(74,144,217,1)',
      borderWidth: 1,
      borderRadius: 3,
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        callbacks: {{
          title: ctx => ctx[0].label,
          label: ctx => ` ${{ctx.raw}} сообщ.`
        }}
      }}
    }},
    scales: {{
      y: {{ beginAtZero: true, ticks: {{ stepSize: 1 }} }},
      x: {{ ticks: {{ maxRotation: 45 }} }}
    }}
  }}
}});

function setMode(mode, btn) {{
  document.querySelectorAll('.btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  chart.data.labels = DATA[mode].labels;
  chart.data.datasets[0].data = DATA[mode].values;
  chart.update();
}}
</script>
</body>
</html>
"""
    return html


def generate_excel(messages):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Устанавливаю openpyxl...")
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аренда мотобайков"

    headers = ["Дата", "Канал", "Текст сообщения", "Ссылка"]
    header_fill = PatternFill("solid", fgColor="4A90D9")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for row_idx, msg in enumerate(messages, 2):
        dt = parse_date(msg["date"])
        date_str = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        channel = msg["title"] or msg["username"] or ""
        text = (msg["text"] or "").replace("\n", " ").strip()
        link = build_message_link(msg["username"], msg["message_id"])

        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=channel)
        ws.cell(row=row_idx, column=3, value=text)

        link_cell = ws.cell(row=row_idx, column=4, value=link if link else "—")
        if link:
            link_cell.hyperlink = link
            link_cell.font = Font(color="1155CC", underline="single")

    # Автоширина
    col_widths = [18, 22, 80, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Перенос текста для колонки "Текст"
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path = REPORTS_DIR / "moto_rental_messages.xlsx"
    wb.save(out_path)
    return out_path


def main():
    print(f"Подключаюсь к БД: {DB_PATH}")
    messages = fetch_messages()
    print(f"Найдено сообщений: {len(messages)}")

    # HTML-график
    html = generate_chart_html(messages)
    html_path = REPORTS_DIR / "moto_rental_chart.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"HTML-график: {html_path}")

    # Excel
    xlsx_path = generate_excel(messages)
    print(f"Excel-выгрузка: {xlsx_path}")

    print("Готово.")


if __name__ == "__main__":
    main()
